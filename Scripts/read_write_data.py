import numpy as np
import pandas as pd
import openpyxl
import xlsxwriter

#region CSV FILES
# READ FROM CSV
print('-- READING CSV --')
# If we omit delimiter parameter, by default ','
print(pd.read_csv('../files/data.csv', header=0, delimiter=';')) # Header_row=0
print(pd.read_csv('../files/data.csv', header=0, nrows=2, delimiter=';')) # Only 2 first rows
print(pd.read_csv('../files/data.csv', header=None, delimiter=';')) # Header = 0,1,2,3,... row=0 is considered data

data_frame_csv1 = pd.DataFrame(pd.read_csv('../files/data.csv', header=0, delimiter=';'))

print(data_frame_csv1[-2:][0:2]) # (data_frame[-2:])[1:2] -> df[-2:]=df1, df1[0:2]=df2 NOT COLUMNS of df1!! (df.loc()!)
print(data_frame_csv1.loc(0)) # _LocIndexer object
print(data_frame_csv1.loc[0]) # row 0 (whole columns)
print(data_frame_csv1['COL2'].loc[0])
#print(data_frame_csv1[1].loc[0]) # It doesn't work! df['key'] key=str

# WRITE TO CSV
print('\n-- WRITTING CSV --\n... :)')
"""
- encoding='utf-8' -> to avoid getUnicodeEncodeError 
- Header/Index - by default True
- sep -> by default ','
- na_rep -> value for missing values (f.e. for NaN values na_rep=0)
- columns -> array with names of columns to write to CSV
- mode -> by default 'w' (overwrite existing file), 'a' to append keeping existing data
"""
data_frame_csv1.to_csv('../files/written.csv', sep='/', header=True, index=False, na_rep='missing', encoding='utf-8')
data_frame_csv1.to_csv('../files/writtenCOL3COL1.csv', index=False, columns=['COL3','COL1'])
data_frame_csv1.to_csv('../files/writtenCOL3COL1.csv', header=False, index=False, columns=['COL3','COL1'], mode='a')


def write_df_to_csv(file_path, data_frame, sep=",", mode='a', **to_csv_kwargs):
    """
    This method write on [file_path] CSV file data of [data_frame] DataFrame, using [sep] as
        char separator.
    If CSV file doesn't exist yet, this method will create it.
    If CSV file already exist, this method will add new data at the end unless [overwrite_file]
        is specified to be True.
    :param file_path: path of the CSV to write in, if doesn't exist this method will create it
    :param data_frame: DataFrame with data to write into CSV file
    :param sep: separator, by default ','
    :param mode: by default 'a', if CSV file already exist new data is appended at the end.
                 To overwrite existing file use 'w'.
    :return: None
    """
    import os.path
    # If CSV file already exist and want to append
    if os.path.isfile(file_path) and mode == 'a':
        try:
            csv_file = pd.read_csv(file_path, nrows=1, sep=sep)
            # Checking columns match
            if len(data_frame.columns) != len(csv_file.columns):
                # Different number of columns
                print('EX1')
                raise Exception("Columns don't match!! Dataframe has " + str(len(data_frame.columns)) + " columns. CSV file has " + str(len(csv_file.columns)) + " columns.")
            elif not (data_frame.columns == csv_file.columns).all():
                # Same number of columns, but not the same column names or order
                print('EX2')
                raise Exception("Columns or column order of DataFrame and CSV don't match!!")
            # Not header to append
            to_csv_kwargs['header'] = False
        except Exception:
            return # Write operation not posible

    # index and header -> by default only header when overwritting
    if 'index' not in to_csv_kwargs:
        to_csv_kwargs['index'] = False
    if 'header' not in to_csv_kwargs:
        to_csv_kwargs['header'] = (mode == 'w')

    # Write data from DF to CSV
    data_frame.to_csv(file_path, mode=mode, sep=sep, **to_csv_kwargs)


# TESTING write_df_to_csv()
write_df_to_csv('../files/writtenNew.csv', data_frame_csv1, mode='w', columns=['COL2'], )


#endregion


#region EXCEL FILES
# READ FROM EXCEL
print('\n-- READING EXCEL --')
# Dependecies -> pip install xlrd or conda install -c anaconda xlrd (inside project folder, with activated environment)
file_xlsx = pd.ExcelFile('../files/data.xlsx')
print(file_xlsx.sheet_names)
print(pd.read_excel('../files/data.xlsx', sheet_name='Sheet1')) # Not shapely
print(pd.read_excel(file_xlsx, sheet_name='Sheet2')) # Shapely
data_frame_excel1 = pd.DataFrame(pd.read_excel(file_xlsx, sheet_name='Sheet2'))

# WRITE TO EXCEL
data_frame_excel2 = pd.DataFrame({'C1':['*1','*2'],
                                  'C2':['+1','+2'],
                                  'C3':['-1','-2']})

# Creating new file / Overwritting existing one
data_frame_excel2.to_excel('../files/newFile.xlsx', sheet_name='Sheet3', index=False)

# Adding to existing file / Creating new one if file_path doesn't exist)
def write_df_to_excel(file_path, data_frame, sheet_name='Sheet1', start_row=None,
                      overwrite_sheet=False,
                      **to_excel_kwargs):
    """
    Append data from a DataFrame [df] to existing Excel file [filename] into [sheet_name] Sheet.
    If [sheet_name] doesn't exist in [filename], then this function will create it.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      file_path : File path or existing ExcelWriter
      data_frame : DataFrame whose content will be written to excel file
      sheet_name : Name of sheet in which DataFrame data will be written (default: 'Sheet1')
      start_row : Upper left cell row to dump data frame.
                  By default (startrow=None) append after last row in existing sheets,
                  and at first row (startrow=0) in new sheets
      overwrite_sheet : Overwrite [sheet_name] existing data with new DataFrame data.
      to_excel_kwargs : Arguments which will be passed to `DataFrame.to_excel() (can be dictionary)

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(file_path, engine='openpyxl')

    try:
        # open workbook from existing file
        writer.book = load_workbook(file_path)

        # start_row -> by default get last row if sheet actually exists
        if start_row is None and sheet_name in writer.book.sheetnames and not overwrite_sheet:
            start_row = writer.book[sheet_name].max_row

        # overwrite sheet -> remove existing and create an empty one with same name and 'position' (index) in file
        if overwrite_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets (to 'replicate' original file) -> {title:sheet}
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass # file will be automatically created

    if start_row is None:
        start_row = 0

    # indexes and column header -> by default only headers when truncating
    if 'index' not in to_excel_kwargs:
        to_excel_kwargs['index'] = False
    if 'header' not in to_excel_kwargs:
        to_excel_kwargs['header'] = overwrite_sheet

    # write new data from DF
    data_frame.to_excel(writer, sheet_name, startrow=start_row, **to_excel_kwargs)

    # save the workbook (which contains all original sheets plus new data and modifications)
    writer.save()


# TESTING write_df_to_excel()
# New file
write_df_to_excel('../files/datatest.xlsx', data_frame_excel2, sheet_name='MyTestSheet', header=True)
# New sheet in existing file
write_df_to_excel('../files/datatest.xlsx', data_frame_excel2, sheet_name='MyTestSheet2', header=True)
#endregion


